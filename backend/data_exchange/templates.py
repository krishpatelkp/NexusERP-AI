"""
==========================================================
NexusERP-AI  |  data_exchange  |  templates.py
==========================================================

Downloadable import template generator.

Responsibilities:
  - Generate .xlsx templates with:
      Sheet 1: Data entry sheet (headers + example row)
      Sheet 2: Instructions sheet
      Sheet 3: Valid choices reference
  - Styled for clarity (bold headers, frozen panes,
    column widths, colour-coded header row)

Rules:
  - NO database access
  - NO business logic
  - Pure file generation only
==========================================================
"""

import io

from .constants import (
    ImportModule,
    MODULE_COLUMNS,
    EXAMPLE_ROWS,
    TEMPLATE_FILENAMES,
    CHOICE_HINTS,
)


# ==========================================================
# COLOUR PALETTE
# ==========================================================

HEADER_FILL   = "1F3864"   # Dark navy blue
HEADER_FONT   = "FFFFFF"   # White text
REQUIRED_FILL = "FFE699"   # Yellow — required field indicator
EXAMPLE_FILL  = "E2EFDA"   # Light green — example row
ALT_ROW_FILL  = "F2F2F2"   # Light grey — alternate instruction rows


# ==========================================================
# TEMPLATE GENERATOR
# ==========================================================

class ImportTemplateGenerator:
    """
    Generates a styled Excel template for a given module.

    Usage:
        generator = ImportTemplateGenerator(module="employees")
        buffer    = generator.generate()
        # buffer is a BytesIO ready for HttpResponse
    """

    def __init__(self, module):
        if module not in ImportModule.ALL:
            raise ValueError(f"Unknown module '{module}'.")
        self.module   = module
        self.filename = TEMPLATE_FILENAMES[module]

    def generate(self):
        """
        Generate the template and return a BytesIO buffer.

        Returns:
            io.BytesIO: Excel file buffer.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment,
                Border, Side,
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError(
                "openpyxl is required for template generation. "
                "Run: pip install openpyxl"
            )

        wb = openpyxl.Workbook()

        self._build_data_sheet(wb, openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter)
        self._build_instructions_sheet(wb, openpyxl, Font, PatternFill, Alignment, Border, Side)
        self._build_choices_sheet(wb, openpyxl, Font, PatternFill, Alignment)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ──────────────────────────────────────
    # SHEET 1: DATA ENTRY
    # ──────────────────────────────────────

    def _build_data_sheet(self, wb, opx, Font, Fill, Align, Border, Side, col_letter):
        ws = wb.active
        ws.title = "Import Data"

        all_columns  = MODULE_COLUMNS[self.module]["all"]
        req_columns  = set(MODULE_COLUMNS[self.module]["required"])
        example_row  = EXAMPLE_ROWS[self.module]

        thin_border = Border(
            left   = Side(style="thin"),
            right  = Side(style="thin"),
            top    = Side(style="thin"),
            bottom = Side(style="thin"),
        )

        # Write headers
        for col_idx, col_name in enumerate(all_columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)

            # Style header
            cell.font      = Font(bold=True, color=HEADER_FONT, size=11)
            cell.fill      = Fill(fill_type="solid", fgColor=HEADER_FILL)
            cell.alignment = Align(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin_border

            # Yellow tint for required columns
            if col_name in req_columns:
                cell.fill = Fill(fill_type="solid", fgColor="C00000")  # Red — required

        # Write example row
        for col_idx, col_name in enumerate(all_columns, start=1):
            value = example_row.get(col_name, "")
            cell  = ws.cell(row=2, column=col_idx, value=value)
            cell.fill      = Fill(fill_type="solid", fgColor=EXAMPLE_FILL)
            cell.alignment = Align(horizontal="left", vertical="center")
            cell.font      = Font(italic=True, size=10, color="555555")
            cell.border    = thin_border

        # Set column widths
        for col_idx, col_name in enumerate(all_columns, start=1):
            letter = col_letter(col_idx)
            width  = max(len(col_name) + 4, 16)
            ws.column_dimensions[letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Row height
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 22

    # ──────────────────────────────────────
    # SHEET 2: INSTRUCTIONS
    # ──────────────────────────────────────

    def _build_instructions_sheet(self, wb, opx, Font, Fill, Align, Border, Side):
        ws = wb.create_sheet("Instructions")

        all_columns = MODULE_COLUMNS[self.module]["all"]
        req_columns = set(MODULE_COLUMNS[self.module]["required"])

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 60

        # Title
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value     = f"NexusERP-AI — {self.module.title()} Import Instructions"
        title_cell.font      = Font(bold=True, size=14, color=HEADER_FONT)
        title_cell.fill      = Fill(fill_type="solid", fgColor=HEADER_FILL)
        title_cell.alignment = Align(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # Sub-header row
        headers = ["Column Name", "Required?", "Instructions"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font      = Font(bold=True, color=HEADER_FONT)
            cell.fill      = Fill(fill_type="solid", fgColor="3A3A5C")
            cell.alignment = Align(horizontal="center", vertical="center")

        # One row per column
        instructions = self._get_instructions()
        for row_idx, col_name in enumerate(all_columns, start=3):
            is_req = col_name in req_columns
            note   = instructions.get(col_name, "Enter the value for this field.")

            fill_color = REQUIRED_FILL if is_req else "FFFFFF"
            if row_idx % 2 == 0:
                fill_color = ALT_ROW_FILL if not is_req else REQUIRED_FILL

            for col_idx, value in enumerate([
                col_name,
                "YES ✓" if is_req else "Optional",
                note,
            ], start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = Fill(fill_type="solid", fgColor=fill_color)
                cell.alignment = Align(horizontal="left", vertical="top", wrap_text=True)
                cell.font      = Font(size=10, bold=(col_idx == 1))
                ws.row_dimensions[row_idx].height = 35

    # ──────────────────────────────────────
    # SHEET 3: VALID CHOICES
    # ──────────────────────────────────────

    def _build_choices_sheet(self, wb, opx, Font, Fill, Align):
        ws = wb.create_sheet("Valid Choices")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 50

        ws.merge_cells("A1:B1")
        title = ws["A1"]
        title.value     = "Valid Values Reference"
        title.font      = Font(bold=True, size=13, color=HEADER_FONT)
        title.fill      = Fill(fill_type="solid", fgColor=HEADER_FILL)
        title.alignment = Align(horizontal="center", vertical="center")

        row = 2
        relevant_hints = self._get_relevant_choices()

        for field, choices in relevant_hints.items():
            ws.cell(row=row, column=1, value=field).font  = Font(bold=True, size=11)
            ws.cell(row=row, column=2, value=", ".join(choices)).alignment = Align(wrap_text=True)
            ws.row_dimensions[row].height = 20
            row += 1

    # ──────────────────────────────────────
    # INSTRUCTIONS TEXT PER MODULE
    # ──────────────────────────────────────

    def _get_instructions(self):
        """Return column-specific instruction strings for this module."""
        base = {
            "employee_id":       "Existing Employee ID (e.g. EMP000001). Must already exist in the system.",
            "first_name":        "Employee's first name. Max 100 characters.",
            "middle_name":       "Optional middle name.",
            "last_name":         "Employee's last name. Max 100 characters.",
            "email":             "Unique work email address. Must not already exist.",
            "phone":             "Phone number (digits only, 9–15 digits, optionally starting with +).",
            "alternate_phone":   "Optional alternate phone number.",
            "gender":            f"One of: {', '.join(CHOICE_HINTS['gender'])}",
            "date_of_birth":     "Date format: YYYY-MM-DD (e.g. 1990-05-15).",
            "marital_status":    f"One of: {', '.join(CHOICE_HINTS['marital_status'])}",
            "blood_group":       f"Optional. One of: {', '.join(CHOICE_HINTS['blood_group'])}",
            "department":        "Must match an existing Department name in your company exactly.",
            "designation":       "Must match an existing Designation name in your company exactly.",
            "employment_type":   f"One of: {', '.join(CHOICE_HINTS['employment_type'])}",
            "employee_status":   f"One of: {', '.join(CHOICE_HINTS['employee_status'])}. Default: Probation.",
            "joining_date":      "Date format: YYYY-MM-DD.",
            "confirmation_date": "Optional. Date format: YYYY-MM-DD.",
            "basic_salary":      "Monthly basic salary (number only, no currency symbols). Must be ≥ 0.",
            "date":              "Attendance date. Format: YYYY-MM-DD.",
            "check_in":          "Check-in time. Format: HH:MM (24-hour, e.g. 09:00).",
            "check_out":         "Check-out time. Format: HH:MM (24-hour, e.g. 18:00).",
            "status":            "Attendance/Asset status. See 'Valid Choices' sheet.",
            "remarks":           "Optional free-text remarks.",
            "leave_type":        "Must match an existing Leave Type name in your company exactly.",
            "start_date":        "Date format: YYYY-MM-DD.",
            "end_date":          "Date format: YYYY-MM-DD. Must be ≥ start_date.",
            "reason":            "Reason for leave. Cannot be empty.",
            "is_half_day":       "Yes or No.",
            "month":             "Month number (1–12).",
            "year":              "4-digit year (e.g. 2026).",
            "hra":               "House Rent Allowance. Number ≥ 0.",
            "transport_allowance":"Transport/Conveyance allowance. Number ≥ 0.",
            "other_allowance":   "Any other allowance. Number ≥ 0.",
            "pf_deduction":      "Provident Fund deduction. Number ≥ 0.",
            "tax_deduction":     "Income Tax / TDS deduction. Number ≥ 0.",
            "other_deduction":   "Any other deduction. Number ≥ 0.",
            "asset_tag":         "Unique asset tag (e.g. LAP-0001). Must not already exist.",
            "name":              "Asset name / model description (e.g. Dell Latitude 5420).",
            "category":          "Must match an existing Asset Category name in your company exactly.",
            "vendor":            "Optional. Must match an existing Vendor name in your company exactly.",
            "serial_number":     "Manufacturer serial number (optional).",
            "brand":             "Brand name (optional, e.g. Dell, HP).",
            "model":             "Model name (optional, e.g. Latitude 5420).",
            "description":       "Optional free-text description.",
            "purchase_date":     "Optional. Date format: YYYY-MM-DD.",
            "purchase_cost":     "Optional. Purchase cost (number ≥ 0).",
            "warranty_expiry":   "Optional. Date format: YYYY-MM-DD.",
            "location":          "Optional. Physical storage location.",
            "condition":         f"One of: {', '.join(CHOICE_HINTS['condition'])}",
            "invoice_number":    "Optional. Purchase invoice number.",
            "amount":            "Payment amount. Must be > 0.",
            "payment_date":      "Date format: YYYY-MM-DD.",
            "payment_method":    f"One of: {', '.join(CHOICE_HINTS['payment_method'])}",
            "reference_number":  "Optional. Transaction/reference number.",
        }
        return base

    def _get_relevant_choices(self):
        """Return only choice hints relevant to this module's columns."""
        all_cols = set(MODULE_COLUMNS[self.module]["all"])
        result   = {}

        mappings = {
            "gender":            CHOICE_HINTS["gender"],
            "marital_status":    CHOICE_HINTS["marital_status"],
            "blood_group":       CHOICE_HINTS["blood_group"],
            "employment_type":   CHOICE_HINTS["employment_type"],
            "employee_status":   CHOICE_HINTS["employee_status"],
            "status":            (
                CHOICE_HINTS["status_attendance"] if self.module == ImportModule.ATTENDANCE
                else CHOICE_HINTS["status_inventory"] if self.module == ImportModule.INVENTORY
                else []
            ),
            "is_half_day":       CHOICE_HINTS["is_half_day"],
            "condition":         CHOICE_HINTS["condition"],
            "payment_method":    CHOICE_HINTS["payment_method"],
        }

        for field, choices in mappings.items():
            if field in all_cols and choices:
                result[field] = choices

        return result
