"""
==========================================================
NexusERP-AI  |  data_exchange  |  exporters.py
==========================================================

File-generation layer.

Responsibilities:
  - ExcelExporter: generate .xlsx from queryset/list
  - CSVExporter:   generate .csv from queryset/list
  - PDFExporter:   generate .pdf summary from data

Rules:
  - NO business logic
  - NO database access (data is passed in)
  - NO direct model imports
  - Accepts plain dicts or serializer data
==========================================================
"""

import csv
import io
from datetime import date, datetime


# ==========================================================
# BASE EXPORTER
# ==========================================================

class BaseExporter:
    """
    Abstract base for all exporters.
    Subclasses implement generate(data, columns, title).
    """

    content_type = None
    extension    = None

    def generate(self, data, columns, title="Export"):
        """
        Generate a file buffer from data.

        Args:
            data    (list[dict]):  Rows to export.
            columns (list[str]):   Column names (in order).
            title   (str):         Report title.

        Returns:
            io.BytesIO or io.StringIO buffer.
        """
        raise NotImplementedError


# ==========================================================
# EXCEL EXPORTER
# ==========================================================

class ExcelExporter(BaseExporter):
    """
    Generates styled Excel files using openpyxl.

    Features:
      - Bold header row with company colours
      - Frozen first row
      - Auto column widths
      - Alternate row shading
      - Sheet named after the module
    """

    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension    = ".xlsx"

    def generate(self, data, columns, title="Export"):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils  import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl is required. Run: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit

        header_fill  = PatternFill(fill_type="solid", fgColor="1F3864")
        header_font  = Font(bold=True, color="FFFFFF", size=11)
        alt_fill     = PatternFill(fill_type="solid", fgColor="F2F2F2")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align   = Alignment(horizontal="left", vertical="center")
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        # Header row
        for col_idx, col_name in enumerate(columns, start=1):
            cell           = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border

        # Data rows
        for row_idx, row_data in enumerate(data, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name, "")
                if isinstance(value, (date, datetime)):
                    value = value.isoformat()
                cell           = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = left_align
                cell.border    = thin_border
                if fill:
                    cell.fill = fill

        # Auto column widths
        for col_idx, col_name in enumerate(columns, start=1):
            letter = get_column_letter(col_idx)
            max_len = len(col_name)
            for row_data in data[:100]:  # sample first 100 rows for width
                cell_val = str(row_data.get(col_name, ""))
                max_len  = max(max_len, len(cell_val))
            ws.column_dimensions[letter].width = min(max_len + 4, 50)

        ws.freeze_panes  = "A2"
        ws.row_dimensions[1].height = 28

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# ==========================================================
# CSV EXPORTER
# ==========================================================

class CSVExporter(BaseExporter):
    """
    Generates UTF-8 CSV files.
    Returns BytesIO with BOM for Excel compatibility.
    """

    content_type = "text/csv; charset=utf-8-sig"
    extension    = ".csv"

    def generate(self, data, columns, title="Export"):
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=columns,
            extrasaction="ignore",
            lineterminator="\n",
        )
        writer.writeheader()

        for row in data:
            safe_row = {}
            for col in columns:
                val = row.get(col, "")
                if isinstance(val, (date, datetime)):
                    val = val.isoformat()
                safe_row[col] = val
            writer.writerow(safe_row)

        # Return as BytesIO with UTF-8 BOM (for Excel compatibility)
        buffer = io.BytesIO()
        buffer.write(b"\xef\xbb\xbf")  # UTF-8 BOM
        buffer.write(output.getvalue().encode("utf-8"))
        buffer.seek(0)
        return buffer


# ==========================================================
# PDF EXPORTER
# ==========================================================

class PDFExporter(BaseExporter):
    """
    Generates PDF reports using ReportLab.

    Features:
      - Company header
      - Title and timestamp
      - Styled table with alternating rows
      - Page numbers in footer
    """

    content_type = "application/pdf"
    extension    = ".pdf"

    def generate(self, data, columns, title="Export"):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib           import colors
            from reportlab.lib.units     import cm
            from reportlab.platypus      import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer,
            )
            from reportlab.lib.styles    import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums     import TA_CENTER, TA_LEFT
        except ImportError:
            raise RuntimeError(
                "reportlab is required for PDF export. "
                "Run: pip install reportlab"
            )

        buffer     = io.BytesIO()
        page_size  = landscape(A4) if len(columns) > 6 else A4
        doc        = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm,
        )

        styles  = getSampleStyleSheet()
        story   = []

        # Title
        title_style = ParagraphStyle(
            "Title",
            parent    = styles["Heading1"],
            fontSize  = 16,
            textColor = colors.HexColor("#1F3864"),
            alignment = TA_CENTER,
            spaceAfter= 6,
        )
        subtitle_style = ParagraphStyle(
            "Subtitle",
            parent    = styles["Normal"],
            fontSize  = 9,
            textColor = colors.grey,
            alignment = TA_CENTER,
            spaceAfter= 12,
        )

        story.append(Paragraph(f"NexusERP-AI — {title}", title_style))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')} | "
            f"Total Records: {len(data)}",
            subtitle_style,
        ))
        story.append(Spacer(1, 0.3*cm))

        if not data:
            story.append(Paragraph("No data to display.", styles["Normal"]))
        else:
            # Build table data
            header_row = [
                Paragraph(f"<b>{col}</b>", ParagraphStyle(
                    "TH", parent=styles["Normal"],
                    fontSize=8, textColor=colors.white,
                ))
                for col in columns
            ]
            table_data = [header_row]

            for row in data:
                table_data.append([
                    Paragraph(str(row.get(col, "") or ""), ParagraphStyle(
                        "TD", parent=styles["Normal"], fontSize=8,
                    ))
                    for col in columns
                ])

            # Calculate column widths
            available_width = (page_size[0] - 2*cm)
            col_width = available_width / len(columns)
            col_widths = [col_width] * len(columns)

            table = Table(table_data, colWidths=col_widths, repeatRows=1)

            # Table styles
            style = TableStyle([
                ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#1F3864")),
                ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
                ("FONTSIZE",    (0, 0), (-1, 0), 9),
                ("FONTSIZE",    (0, 1), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                    [colors.white, colors.HexColor("#F2F2F2")]),
                ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("VALIGN",      (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING",(0, 0), (-1, -1), 4),
                ("TOPPADDING",  (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
            ])
            table.setStyle(style)
            story.append(table)

        doc.build(story)
        buffer.seek(0)
        return buffer


# ==========================================================
# EXPORTER FACTORY
# ==========================================================

EXPORTERS = {
    "excel": ExcelExporter,
    "csv":   CSVExporter,
    "pdf":   PDFExporter,
}


def get_exporter(export_format):
    """
    Return the correct exporter for the given format.

    Args:
        export_format (str): 'excel', 'csv', or 'pdf'.

    Returns:
        BaseExporter subclass instance.
    """
    exporter_class = EXPORTERS.get(export_format.lower())
    if not exporter_class:
        from .exceptions import UnsupportedExportFormatError
        raise UnsupportedExportFormatError(
            f"Export format '{export_format}' is not supported. "
            "Supported: excel, csv, pdf"
        )
    return exporter_class()
